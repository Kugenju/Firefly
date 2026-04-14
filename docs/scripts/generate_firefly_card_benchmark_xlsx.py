from __future__ import annotations

import json
import re
import struct
from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
CORE = Path.home() / "AppData/Local/Temp/sts2-decompiled-firefly/sts2/MegaCrit/sts2/Core"

FIREFLY_POOL = ROOT / "Scripts/CardPools/FireflyCardPool.cs"
FIREFLY_CARDS = ROOT / "Scripts/Cards"
FIREFLY_LOC = ROOT / "Firefly/localization/zhs/cards.json"

IRON_POOL = CORE / "Models/CardPools/IroncladCardPool.cs"
IRON_CARDS = CORE / "Models/Cards"
STS2_PCK = Path(r"G:\SteamLibrary\steamapps\common\Slay the Spire 2\SlayTheSpire2.pck")
STS2_ZHS_CARDS = "localization/zhs/cards.json"

OUT_XLSX = ROOT / "docs/firefly_card_benchmark_template.xlsx"


def read_utf8(path: Path) -> str:
    return path.read_text(encoding="utf-8-sig")


def extract_pool_classes(pool_text: str) -> list[str]:
    return re.findall(r"ModelDb\.Card<([A-Za-z0-9_]+)>\(\)", pool_text)


def class_to_key(class_name: str) -> str:
    s1 = re.sub(r"(.)([A-Z][a-z]+)", r"\1_\2", class_name)
    s2 = re.sub(r"([a-z0-9])([A-Z])", r"\1_\2", s1)
    return s2.upper()


def parse_card_meta(card_file: Path) -> dict[str, str]:
    text = read_utf8(card_file)
    m = re.search(r":\s*base\((.*?)\)\s*\{", text, flags=re.S)
    cost = card_type = rarity = target = ""
    if m:
        flat = re.sub(r"\s+", " ", m.group(1))
        parts = [p.strip() for p in flat.split(",")]
        if parts:
            cost = parts[0]
        t = re.search(r"CardType\.([A-Za-z_]+)", flat)
        if t:
            card_type = t.group(1)
        r = re.search(r"CardRarity\.([A-Za-z_]+)", flat)
        if r:
            rarity = r.group(1)
        tg = re.search(r"TargetType\.([A-Za-z_]+)", flat)
        if tg:
            target = tg.group(1)

    return {
        "cost": cost,
        "type": card_type or "Unknown",
        "rarity": rarity or "Unknown",
        "target": target or "",
    }


def _read_u32(f) -> int:
    return struct.unpack("<I", f.read(4))[0]


def _read_u64(f) -> int:
    return struct.unpack("<Q", f.read(8))[0]


def load_pck_file_bytes(pck_path: Path, asset_path: str) -> bytes:
    with pck_path.open("rb") as f:
        magic = f.read(4)
        if magic != b"GDPC":
            raise ValueError(f"Unsupported pck magic: {magic!r}")

        f.read(20)  # pack/godot versions + flags
        file_base = _read_u64(f)
        dir_offset = _read_u64(f)

        f.seek(dir_offset)
        entry_count = _read_u32(f)

        target_offset: int | None = None
        target_size: int | None = None
        for _ in range(entry_count):
            name_len = _read_u32(f)
            name = f.read(name_len).decode("utf-8", errors="replace").rstrip("\x00")
            pad = (-name_len) % 4
            if pad:
                f.read(pad)

            entry_offset = _read_u64(f)
            entry_size = _read_u64(f)
            f.read(16)  # md5
            flags = _read_u32(f)

            if name == asset_path:
                if flags != 0:
                    raise ValueError(f"Unsupported pck entry flags for {asset_path}: {flags}")
                target_offset = entry_offset
                target_size = entry_size
                break

        if target_offset is None or target_size is None:
            raise FileNotFoundError(f"Asset not found in pck: {asset_path}")

        f.seek(target_offset + file_base)
        return f.read(target_size)


def load_sts2_zhs_cards() -> dict[str, str]:
    if not STS2_PCK.exists():
        return {}
    try:
        payload = load_pck_file_bytes(STS2_PCK, STS2_ZHS_CARDS)
        return json.loads(payload.decode("utf-8-sig"))
    except Exception:
        return {}


def build_firefly_rows() -> list[dict[str, str]]:
    classes = extract_pool_classes(read_utf8(FIREFLY_POOL))
    loc = json.loads(read_utf8(FIREFLY_LOC))

    rows: list[dict[str, str]] = []
    for class_name in classes:
        f = FIREFLY_CARDS / f"{class_name}.cs"
        meta = parse_card_meta(f)
        key = class_to_key(class_name)
        rows.append(
            {
                "class": class_name,
                "name_zh": loc.get(f"{key}.title", ""),
                "key": key,
                "desc": loc.get(f"{key}.description", ""),
                "cost": meta["cost"],
                "type": meta["type"],
                "rarity": meta["rarity"],
                "target": meta["target"],
                "file": str(f.relative_to(ROOT)),
            }
        )
    return rows


def build_iron_rows() -> list[dict[str, str]]:
    classes = extract_pool_classes(read_utf8(IRON_POOL))
    loc = load_sts2_zhs_cards()
    rows: list[dict[str, str]] = []
    for class_name in classes:
        f = IRON_CARDS / f"{class_name}.cs"
        if not f.exists():
            continue
        meta = parse_card_meta(f)
        key = class_to_key(class_name)
        rows.append(
            {
                "class": class_name,
                "name_zh": loc.get(f"{key}.title", ""),
                "desc_zh": loc.get(f"{key}.description", ""),
                "cost": meta["cost"],
                "type": meta["type"],
                "rarity": meta["rarity"],
                "target": meta["target"],
            }
        )
    return rows


def style_sheet(ws, wrap_text: bool, width_map: list[int]) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c in range(1, ws.max_column + 1):
        cell = ws.cell(1, c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            cell.border = border
            if wrap_text:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    for i, w in enumerate(width_map, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def main() -> None:
    firefly_rows = build_firefly_rows()
    iron_rows = build_iron_rows()

    fw_total, ir_total = len(firefly_rows), len(iron_rows)
    fw_type = Counter(r["type"] for r in firefly_rows)
    ir_type = Counter(r["type"] for r in iron_rows)
    fw_rarity = Counter(r["rarity"] for r in firefly_rows)
    ir_rarity = Counter(r["rarity"] for r in iron_rows)
    fw_matrix = Counter((r["type"], r["rarity"]) for r in firefly_rows)
    ir_matrix = Counter((r["type"], r["rarity"]) for r in iron_rows)

    all_types = sorted(set(fw_type) | set(ir_type))
    all_rarities = sorted(set(fw_rarity) | set(ir_rarity))

    placeholder_rows: list[list[str]] = []
    for card_type in all_types:
        for rarity in all_rarities:
            gap = ir_matrix.get((card_type, rarity), 0) - fw_matrix.get((card_type, rarity), 0)
            for i in range(1, max(gap, 0) + 1):
                placeholder_rows.append(
                    [
                        f"{card_type[:3].upper()}_{rarity[:3].upper()}_{i:02d}",
                        card_type,
                        rarity,
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                    ]
                )

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "current_cards"
    ws1.append(
        [
            "\u5e8f\u53f7",
            "Class",
            "\u4e2d\u6587\u540d",
            "\u7c7b\u578b",
            "\u7a00\u6709\u5ea6",
            "\u8d39\u7528",
            "\u76ee\u6807",
            "\u672c\u5730\u5316Key",
            "\u63cf\u8ff0",
            "\u6587\u4ef6",
        ]
    )
    for i, r in enumerate(firefly_rows, 1):
        ws1.append(
            [
                i,
                r["class"],
                r["name_zh"],
                r["type"],
                r["rarity"],
                r["cost"],
                r["target"],
                r["key"],
                r["desc"],
                r["file"],
            ]
        )

    ws2 = wb.create_sheet("benchmark")
    ws2.append(
        [
            "\u7ef4\u5ea6",
            "Firefly\u5f53\u524d",
            "Ironclad\u57fa\u51c6",
            "\u5dee\u503c(\u57fa\u51c6-\u5f53\u524d)",
            "\u5907\u6ce8",
        ]
    )
    ws2.append(
        [
            "\u603b\u5361\u724c\u6570",
            fw_total,
            ir_total,
            ir_total - fw_total,
            "\u57fa\u4e8e IroncladCardPool \u5b9e\u9645\u5361\u6c60",
        ]
    )
    ws2.append([])
    ws2.append(["\u6309\u7c7b\u578b", "", "", "", ""])
    for card_type in all_types:
        ws2.append([card_type, fw_type.get(card_type, 0), ir_type.get(card_type, 0), ir_type.get(card_type, 0) - fw_type.get(card_type, 0), ""])
    ws2.append([])
    ws2.append(["\u6309\u7a00\u6709\u5ea6", "", "", "", ""])
    for rarity in all_rarities:
        ws2.append([rarity, fw_rarity.get(rarity, 0), ir_rarity.get(rarity, 0), ir_rarity.get(rarity, 0) - fw_rarity.get(rarity, 0), ""])

    ws3 = wb.create_sheet("gap_matrix")
    ws3.append(["\u7c7b\u578b", "\u7a00\u6709\u5ea6", "Firefly\u5f53\u524d", "Ironclad\u57fa\u51c6", "\u7f3a\u53e3"])
    for card_type in all_types:
        for rarity in all_rarities:
            fwv = fw_matrix.get((card_type, rarity), 0)
            irv = ir_matrix.get((card_type, rarity), 0)
            ws3.append([card_type, rarity, fwv, irv, irv - fwv])

    ws_iron = wb.create_sheet("ironclad_cards")
    ws_iron.append(
        [
            "\u5e8f\u53f7",
            "Class",
            "\u4e2d\u6587\u540d",
            "\u4e2d\u6587\u63cf\u8ff0",
            "\u7c7b\u578b",
            "\u7a00\u6709\u5ea6",
            "\u8d39\u7528",
            "\u76ee\u6807",
        ]
    )
    for i, row in enumerate(iron_rows, 1):
        ws_iron.append(
            [
                i,
                row["class"],
                row["name_zh"],
                row["desc_zh"],
                row["type"],
                row["rarity"],
                row["cost"],
                row["target"],
            ]
        )

    ws4 = wb.create_sheet("placeholders")
    ws4.append(
        [
            "\u5360\u4f4dID",
            "\u5efa\u8bae\u7c7b\u578b",
            "\u5efa\u8bae\u7a00\u6709\u5ea6",
            "\u8d39\u7528(\u5f85\u586b)",
            "Class\u540d(\u5f85\u586b)",
            "\u4e2d\u6587\u540d(\u5f85\u586b)",
            "\u6548\u679c\u63cf\u8ff0(\u5f85\u586b)",
            "\u673a\u5236\u5f52\u5c5e(\u707c\u70ed/\u8424\u706b/\u88c2\u89e3/\u751f\u5b58)",
            "\u5907\u6ce8",
        ]
    )
    for row in placeholder_rows:
        ws4.append(row)
    if not placeholder_rows:
        ws4.append(["(\u65e0\u7f3a\u53e3)", "", "", "", "", "", "", "", ""])

    ws5 = wb.create_sheet("readme")
    ws5.append(["\u9879", "\u5185\u5bb9"])
    ws5.append(["\u751f\u6210\u65f6\u95f4", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws5.append(["Firefly\u5361\u6e90", str(FIREFLY_POOL.relative_to(ROOT))])
    ws5.append(["Ironclad\u5361\u6e90", str(IRON_POOL)])
    ws5.append(
        [
            "\u8bf4\u660e",
            "\u5360\u4f4d\u6a21\u677f\u6309 Ironclad \u7684 \u7c7b\u578b\u00d7\u7a00\u6709\u5ea6 \u5206\u5e03\u4e0e\u5f53\u524d Firefly \u5b9e\u9645\u5361\u6c60\u5dee\u503c\u81ea\u52a8\u751f\u6210\u3002",
        ]
    )

    style_sheet(ws1, wrap_text=True, width_map=[6, 30, 18, 12, 12, 10, 14, 26, 56, 40])
    style_sheet(ws2, wrap_text=False, width_map=[18, 14, 14, 14, 34])
    style_sheet(ws3, wrap_text=False, width_map=[14, 14, 14, 14, 10])
    style_sheet(ws_iron, wrap_text=True, width_map=[6, 34, 18, 64, 14, 14, 10, 14])
    style_sheet(ws4, wrap_text=True, width_map=[16, 12, 12, 12, 24, 20, 48, 30, 28])
    style_sheet(ws5, wrap_text=True, width_map=[18, 100])

    sub_fill = PatternFill("solid", fgColor="D9E1F2")
    for r in range(1, ws2.max_row + 1):
        if ws2.cell(r, 1).value in ("\u6309\u7c7b\u578b", "\u6309\u7a00\u6709\u5ea6"):
            for c in range(1, ws2.max_column + 1):
                ws2.cell(r, c).fill = sub_fill
                ws2.cell(r, c).font = Font(bold=True)

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)

    print(str(OUT_XLSX))
    print(f"Firefly={fw_total}, Ironclad={ir_total}, placeholders={len(placeholder_rows)}")


if __name__ == "__main__":
    main()
